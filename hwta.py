import logging
import sys
import os
import re
import openpyxl
import json
from configparser import ConfigParser

class ReadConfig:
    def __init__(self, filepath=None):
        if filepath:
            configPath = filepath
        else:
            curPath = os.path.dirname(os.path.realpath(sys.argv[0]))
            configPath = os.path.join(curPath, "config.ini")
        self.cf = ConfigParser()
        self.cf.read(configPath, encoding="utf-8-sig")

    def datapath(self, param):
        value = self.cf.get("datapath", param)
        return value

    def mask(self, param):
        value = self.cf.get("mask", param)
        return value

    def question(self, param):
        value = self.cf.get("question", param)
        return value

    def excel(self, param):
        value = self.cf.get("excel", param)
        return value

class CustomFormatter(logging.Formatter):
   '''Logging Formatter to add colors and count warning / errors'''
   grey = '\x1b[38;21m'
   yellow = '\x1b[33;21m'
   red = '\x1b[31;21m'
   bold_red = '\x1b[31;1m'
   reset = '\x1b[0m'
   format = '%(asctime)s - %(message)s (%(filename)s:%(lineno)d) - %(name)s - %(levelname)s' 
   
   FORMATS = {
       logging.DEBUG: grey+format+reset,
       logging.INFO: grey+format+reset,
       logging.WARNING: yellow+format+reset,
       logging.ERROR: red+format+reset,
       logging.CRITICAL: bold_red+format+reset
   }

   def format(self, record):
       log_fmt = self.FORMATS.get(record.levelno)
       formatter = logging.Formatter(log_fmt)
       return formatter.format(record)

class Grader():
    def __init__(self, classname,configFile):
        self.base_path = os.getcwd()
        homework_ind = configFile.question("homework_ind")
        self.source_path = configFile.datapath("source_path") + '/' + classname + '/' + homework_ind#let grader know where to find the source homework
        self.bak_path = configFile.datapath("bak_path") + '/' + classname + '/' + homework_ind
        self.paper_path = classname + '/' + configFile.datapath("paper_path") + '/' + homework_ind
        self.question_path = configFile.datapath("question_path")#let grader know where to find the question bank
        self.mark_path = classname + '/' + configFile.datapath("mark_path") + '/' + homework_ind
        self.wrap_path = classname + '/' + configFile.datapath("wrap_path")
        self.xlsx_path = classname + '/' + configFile.datapath("xlsx_path")
        self.question_ind = configFile.question('question_ind')
        self.id_status = self.wrap_path + '/' + 'id_status.json'
        self.class_id = classname
        self.cur_stu_id = '0'
        self.cur_hw_id = homework_ind
        self.hw_status = {}
        self.id_col = configFile.excel('id_col')
        self.score_col = configFile.excel('score_col')
        self.qb_start_pos = int(configFile.question('question_body_start_pos'))
        self.qb_end_pos = int(configFile.question('question_body_end_pos'))
        self.timer_chk = int(configFile.mask('timer'))
    
    def make_fs(self):
        #os.makedirs(self.source_path, exist_ok=True) #preconstructed outside the Grader
        #os.makedirs(self.questions_path, exist_ok=True) #preconstructed outside the Grader
        os.makedirs(self.bak_path, exist_ok=True)
        os.makedirs(self.paper_path, exist_ok=True)
        os.makedirs(self.mark_path, exist_ok=True)
        os.makedirs(self.wrap_path, exist_ok=True)
        os.makedirs(self.xlsx_path, exist_ok=True)
        if not os.path.exists(self.xlsx_path + '/' + self.class_id + '.xlsx'):
            mv_xlsx_cmd = 'cp' + ' ' + '-rf' + ' ' + self.class_id + '.xlsx' + ' ' + self.xlsx_path
            os.system(mv_xlsx_cmd)
        if not os.path.exists(self.wrap_path+'/'+'suffix.py'):
            ln_suffix_cmd = 'ln' + ' ' + os.getcwd() + '/' + self.question_path + '/' + 'suffix.py' + ' ' + self.wrap_path
            os.system(ln_suffix_cmd)
        if os.path.exists(self.id_status):
            with open(self.id_status, 'r') as f:
                self.id_status = json.load(f)
        else:
            self.id_status = {}#eg:{20220901001:[1,3,1],20220901002:[1,4,0],20220901003:[0,0,0]}//{学号:[marked_flag,score,plagiarized_flag]} 
    
    def checked_in_xlsx(self, content, xlsx_file, region):
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
        #if content in [item.value for item in ws[region]]:
        #    return True
        #else:
        #    return False
        ind = [i for i, item in enumerate(ws[region]) if item.value==content]
        if not content in self.id_status and ind:
            self.id_status[content]=[[region, str(ind[0]+1)]]#Notes: 1:'+1' due to excel row is from 1 to start;2:[[]] means the structure is list and the first elment is list
            #self.id_status[content].append(0)#the second position is for marked_flag, initial value is 0: not marked yet
            return True
        elif content in self.id_status:
            return True
        else:
            return False
    
    def fetch_paper(self):
        mod_cmd = 'chmod 700 -R %s' % self.source_path
        #os.system(mod_cmd)
        bak_src_cmd = 'rsync -at %s/*.py %s/' % (self.source_path, self.bak_path)
        os.system(bak_src_cmd)
        s_list = os.listdir(self.source_path)
        #s_list.sort()
        got_paper = False
        hw_ind = int(re.match(r'hw(\d+)',self.cur_hw_id).group(1))
        for s in s_list:
            s_id = re.match(r'(\d{10}).*\.py$', s)#get s_id from paper name
            if not s_id:
                continue
            s_id = s_id.group(1)
            s_id_hw = self.paper_path + '/' + s
            if not os.path.exists(s_id_hw) and self.checked_in_xlsx(s_id, self.xlsx_path+'/'+self.class_id+'.xlsx', self.id_col): #put in self.paper_path if s_id in excel and not in self.paper_path
                mv_src_cmd = 'mv' + ' ' + self.source_path + '/' + s + ' ' + self.paper_path
                os.system(mv_src_cmd)
                #self.id_status[s_id].append(0)  # the second position is for marked_flag, initial value is 0: not marked yet
                while(hw_ind+1>len(self.id_status[s_id])):
                    self.id_status[s_id].append([0])  # the second position is a list for hw1: [marked_flag(initial value is 0: not marked yet)], the list can append score for hw1
                got_paper = True
        return got_paper
                
    def get_assign(self):
        a_dict = {}
        self.hw_status[self.cur_hw_id]={}
        q_str = self.question_ind
        q_list = q_str.split(',')
        q_files = [q+'.py' for q in q_list]
        for q in q_files:
            num = 0
            q_name = self.question_path + '/' + q
            with open (q_name, 'r') as f:
                prefix = f.readlines()
            #a_dict[prefix[self.qb_start_pos]]=[]
            in_comment = False
            in_embeded_code = False
            def_key = None
            for p_prefix in prefix:
                if def_key:
                    a_dict[def_key].append(p_prefix)
                if (len(p_prefix)>4 and p_prefix[0:4] == "def ") or (len(p_prefix)>6 and p_prefix[0:6] == "class "):
                    def_key = p_prefix
                    a_dict[def_key] = []
                    if self.timer_chk != 0 and p_prefix[0:4]=="def ":#if decorating class it seems not right for hw4, need to check further
                        a_dict[def_key].append('@func_set_timeout(%d)\n'%self.timer_chk)
                    a_dict[def_key].append(p_prefix)
                #a_dict[prefix[self.qb_start_pos]].append(p_prefix)
                if in_embeded_code == True:
                    if p_prefix[0:5] == "    #":
                        #a_dict[prefix[self.qb_start_pos]].pop()
                        a_dict[def_key].pop()
                        in_embeded_code=False
                        break
                if in_comment == True:
                    if p_prefix == "    '''\n":
                        in_embeded_code=True
                        in_comment=False
                if p_prefix == "    '''\n":
                    if not in_embeded_code:
                        in_comment = True
            #for p in a_dict[prefix[self.qb_start_pos]]:
            for p in a_dict[def_key]:
                if p[4:8]=='>>> ':
                    num += 1
            q_ind = q.split('.')[0]
            self.hw_status[self.cur_hw_id][q_ind] = num
        return(a_dict)
        
    def marked_checker(self, f): #check if marked already
        stu_id = re.match(r'(\d{10}).*\.py$', f).group(1)
        self.cur_stu_id = stu_id
        #if (not stu_id in self.id_status):
            ##self.id_status[stu_id]=[0]
            #self.id_status[stu_id].append(0)
        #return self.id_status[stu_id][0]
        hw_ind = int(re.match(r'hw(\d+)',self.cur_hw_id).group(1))#homework id is from 1 to start while the index of python is 0 based, BUT
        return self.id_status[stu_id][hw_ind][0]#hw_ind is 1, the postion is right 1 since position 0 is always the (row,col) of stu_id, [0] is the marked_flag position in list positioned by hw_ind
        
    def wrapper(self, paper, items, wrapfile):
        with open(paper, 'r') as fr:
            paper_code = fr.readlines()
        assign_set = set(items.keys())
        with open(wrapfile,'w') as fw:
            if self.timer_chk != 0:
                fw.write('from func_timeout import func_set_timeout\n')
            fw.write('current_student_id=%d\n'%int(self.cur_stu_id))
            for code_line in paper_code:
                code_line_feature = re.match(r'(^[c|d].+\().*\)', code_line)
                if code_line_feature:
                    code_line_feature = code_line_feature.group(1)
                e_addr = re.match(r'^#email:(.+@.+)\n$', code_line)
                if e_addr:
                    e_addr=e_addr.group(1).strip()
                    self.id_status[self.cur_stu_id][0].append(e_addr)#append after the row&col elements of the first element
                if code_line in assign_set:
                    assign_set.remove(code_line)
                    doctest_line = items[code_line]
                    for cl in doctest_line:
                        fw.write(cl)
                else:
                    fw.write(code_line)
                    for assign_set_element in assign_set:
                        if code_line_feature and code_line_feature in assign_set_element:
                            assign_set.remove(assign_set_element)
                            doctest_line = items[assign_set_element][1:]
                            for cl in doctest_line:
                                fw.write(cl)
                            break         
            fw.flush()

    def tester(self):
        print('\r%s'%self.cur_stu_id, end='\r')
        test_path = self.wrap_path
        os.chdir(test_path)
        run_case_cmd = 'python3 suffix.py'
        fh = os.popen(run_case_cmd, 'r')
        buff = fh.readlines()
        fh.close()
        os.chdir(self.base_path)
        return (buff)

    def marker(self, result):
        cp_cmd = 'cp' + ' ' + '-rf' + ' ' + self.paper_path + '/' + self.cur_stu_id + '.py' + ' ' + self.mark_path + '/'
        os.system(cp_cmd)
        #self.id_status[self.cur_stu_id] = [1]
        #self.id_status[self.cur_stu_id][1] = 1
        hw_ind = int(re.match(r'hw(\d+)', self.cur_hw_id).group(1))
        self.id_status[self.cur_stu_id][hw_ind][0] = 1
        stu_marked_paper = self.mark_path + '/' + self.cur_stu_id + '.py'
        write_num = 0
        with open(stu_marked_paper, 'a') as m:
            m.write('\n\n\n')
            m.write('**********Below is the comment region***********\n')
            if not result:
                score = 100.0
                self.id_status[self.cur_stu_id][hw_ind].append(score)
                m.write('Passed all the test cases, score is: %.0f! \n' % score)
                m.write('Congratulations!')
            else:
                for line in result:
                    if(write_num):
                        m.write(line)
                        write_num -= 1
                        continue
                    if line == 'Failed example:\n':
                        m.write(line)
                        write_num = 1
                        continue
                    if line == 'Exception raised:\n':
                        m.write(line)
                        continue
                    if re.match(r'\s+\w+Error.+', line):
                        m.write(line)
                        continue
                    if line == 'Expected:\n':
                        m.write(line)
                        write_num= 1
                        continue
                    if line == 'Got:\n':
                        m.write(line)
                        write_num = 1
                        continue
                    if re.match(r'(\d+) items had failures:\n', line):
                        error_items = re.match(r'(\d+) items had failures:\n', line).group(1)
                        error_items = int(error_items)
                        m.write('\n')
                        m.write('%d items had failures out of %d\n' % (error_items, len(self.hw_status[self.cur_hw_id])))
                        write_num = error_items
                        continue
                    if re.match(r'.+\d+ failures\.\n', line):
                        error_cases = re.match(r'.+?(\d+) failures\.\n', line).group(1)
                        error_cases = int(error_cases)
                        m.write('%d cases had failures out of %d\n' % (error_cases, sum(self.hw_status[self.cur_hw_id].values())))
                        score = round(-error_cases/sum(self.hw_status[self.cur_hw_id].values())*40+100,0)
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('\n')
                        m.write('Score is: %.0f' % score)
                        continue
                    if line == 'SyntaxError\n':
                        m.write(line)
                        m.write('\n')
                        score = 59.0
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('Score is %.0f' % score)
                        continue
                    if line == 'IndentationError\n':
                        m.write(line)
                        m.write('\n')
                        score = 59.0
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('Score is %.0f' % score)
                        continue
                    if line == 'NameError\n':
                        m.write(line)
                        m.write('\n')
                        score = 59.0
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('Score is %.0f' % score)
                        continue
                    if line == 'TypeError\n':
                        m.write(line)
                        m.write('\n')
                        score = 59.0
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('Score is %.0f' % score)
                        continue
                    if line == 'AttributeError\n':
                        m.write(line)
                        m.write('\n')
                        score = 59.0
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('Score is %.0f' % score)
                        continue
                    if re.match(r'^[A-Z][a-z]*Error\n$', line):
                        m.write(line)
                        m.write('\n')
                        score = 59.0
                        self.id_status[self.cur_stu_id][hw_ind].append(score)
                        m.write('Score is %.0f' % score)
                        continue

    def recoder(self):
        with open(self.wrap_path + '/' + 'id_status.json', 'w') as fid:
            json.dump(self.id_status, fid, indent=4, sort_keys=True)
        hw_ind=int(re.match(r'hw(\d+)',self.cur_hw_id).group(1))
        wb = openpyxl.load_workbook(self.xlsx_path+'/'+self.class_id+'.xlsx')
        ws = wb.active
        for sidv in self.id_status.values():
            cr = self.score_col+sidv[0][1]
            if len(sidv)>hw_ind:
                if len(sidv[hw_ind])>1: ws[cr]=sidv[hw_ind][1]
        wb.save(self.xlsx_path+'/'+self.class_id+'.xlsx')

    def post_mail(self):
        import yagmail
        auth_code = sys.argv[1]
        yag = yagmail.SMTP(user='267282100@qq.com', password=auth_code, host='smtp.qq.com')
        marked_papers = os.listdir(self.mark_path)
        for mp in marked_papers:
            stu_id = re.match(r'^(\d+)\.py$',mp).group(1)
            with_addr = self.id_status[stu_id][0]
            if len(with_addr)>2:
                to_addr = with_addr[-1]
                sub = self.cur_hw_id + ' comments'
                stu_marked_paper = self.mark_path + '/' + mp
                yag.send(to=to_addr, subject=sub, contents='comments to your homwork as attached:\n', attachments=stu_marked_paper)

    def liner(self):
        self.make_fs()    
        got_paper = self.fetch_paper()#from source_path to clean into paper_path and copy to bak_path
        if not got_paper:
            return False
        assign_items = self.get_assign()#return dictionary such as {'def square():': content, 'def sqrt():':content, 'def sum():':content}
        p_list = os.listdir(self.paper_path)
        for p in p_list:
            #finished_item_num = 0
            marked = self.marked_checker(p)#check if not test due to late submission, set id_status dict and mkdir hwx
            if not marked:
                paper_file = self.paper_path + '/' + p
                wrap_file = self.wrap_path + '/' + 'test.py'
                self.wrapper(paper_file, assign_items, wrap_file)#input: p, self.questions_path, self_questions_list and self_exec_path. insert docstring and __main__ to generate executable python
                stdout_buff = self.tester()#input: p_wraped, self.marks_path. run and mark operation to generate marked paper
                self.marker(stdout_buff)
        self.recoder()#input: self.id_status and self.excel_path
        #self.post_mail()
        return True
            
    def cleanup(self):
        mod_cmd = 'chmod 775 -R %s' % self.source_path
        os.system(mod_cmd)


if __name__=='__main__':
    #create logger with 'spam_application'
    logger = logging.getLogger('HwTA')     
    #create console handler with a higher log level
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)    
    if len(sys.argv)>1:
        if sys.argv[1] == 'color':
            ch.setFormatter(CustomFormatter())    
    if len(sys.argv)>2:
        logger.setLevel(logging.__getattribute__(sys.argv[2]))
    else:
        logger.setLevel(logging.DEBUG)    
    logger.addHandler(ch)
        
    conf = ReadConfig()
    class_list = conf.datapath('class_path')
    class_list = class_list.split(',')
    for cls in class_list:
        os.makedirs(cls, exist_ok=True)
        graderTA = Grader(cls, conf)
        graderTA.liner()
        print('finished %s class' % cls)
        graderTA.cleanup()#rm source_path file and set source_path rw right to 775
        post_mail_en = conf.mask('post')
        if post_mail_en == '1':
            graderTA.post_mail()
    
